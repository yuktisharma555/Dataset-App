import streamlit as st
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# ---------------- UI ----------------
st.set_page_config(layout="wide")

st.markdown("""
<style>
body {
    background: linear-gradient(135deg, #dbeafe, #ffffff, #ede9fe);
}
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 style='text-align:center;'>📊 Smart Dataset Cleaner Pro</h1>", unsafe_allow_html=True)

file = st.file_uploader("Upload Excel", type=["xlsx"])
dataset_type = st.selectbox("Dataset Type", ["medicine","vaccine","testkit"])

# ---------------- COLORS ----------------
dark_red = PatternFill("solid", fgColor="8B0000")
light_red = PatternFill("solid", fgColor="FF9999")
light_orange = PatternFill("solid", fgColor="FFD580")

# ---------------- MAIN ----------------
if file:
    try:
        df = pd.read_excel(file, header=5)
        df.columns = df.columns.astype(str).str.upper().str.strip()

        original_df = df.copy()

        st.success("✅ Header loaded correctly")

        # COLUMN MAPPING
        desc = "GOODS DESCRIPTION"
        qty = "QUANTITY"
        price = "ITEM PRICE_INV"
        total_price = "TOTAL PRICE_INV_FC"
        country = "COUNTRY"
        exporter = "EXPORTER"
        consignee = "CONSIGNEE NAME"
        std_qty_col = "STD QUANTITY"
        currency_col = "CURRENCY"

        required = [desc, qty, total_price, country]
        missing = [c for c in required if c not in df.columns]

        if missing:
            st.error(f"❌ Missing columns: {missing}")
            st.stop()

        df[qty] = pd.to_numeric(df[qty], errors="coerce")
        df[price] = pd.to_numeric(df.get(price, 0), errors="coerce")
        df[total_price] = pd.to_numeric(df[total_price], errors="coerce")

        # ---------------- USD (UNIT PRICE) ----------------
        exchange_rates = {
            "ZAR": 0.0628, "THB": 0.0322, "CAD": 0.7334, "INR": 0.0109,
            "MXN": 0.058, "RUB": 0.0129, "GBP": 1.3455, "EUR": 1.1821,
            "AED": 0.2722, "USD": 1
        }

        def convert_to_usd(row):
            curr = str(row.get(currency_col, "USD")).upper()
            rate = exchange_rates.get(curr, 1)
            return row[price] * rate if pd.notna(row[price]) else 0

        df["Item Price (USD)"] = df.apply(convert_to_usd, axis=1)

        if currency_col in df.columns:
            cols = list(df.columns)
            cols.remove("Item Price (USD)")
            idx = cols.index(currency_col)
            cols.insert(idx + 1, "Item Price (USD)")
            df = df[cols]

        # ---------------- CLASSIFICATION (RESTORED) ----------------
        def classify(d):
            d = str(d).lower()

            if dataset_type == "vaccine":
                return "Pediatric" if "pediatric" in d else "Adult"

            elif dataset_type == "medicine":
                return "API" if "api" in d else "Formulation"

            elif dataset_type == "testkit":
                if "card" in d:
                    return "Card"
                elif "strip" in d or "dipstick" in d:
                    return "Strip"
                elif "test kit" in d or "test" in d:
                    return "Full Testing Kit"
                else:
                    return "Other"

            return "Other"

        df["Classification"] = df[desc].apply(classify)

        cols = list(df.columns)
        cols.remove("Classification")
        desc_index = cols.index(desc)
        cols.insert(desc_index + 1, "Classification")
        df = df[cols]

        # ---------------- TESTKIT LOGIC ----------------
        if dataset_type == "testkit":

            def extract_number(text):
                text = str(text).lower()
                match = re.search(r'(\d+)', text)
                return int(match.group(1)) if match else 1

            df["Standard Quantity"] = df[desc].apply(extract_number)
            df["Adjusted Quantity"] = df["Standard Quantity"] * df[qty]

            cols = list(df.columns)
            cols.remove("Standard Quantity")
            cols.remove("Adjusted Quantity")

            q_idx = cols.index(qty)

            cols.insert(q_idx, "Standard Quantity")
            cols.insert(q_idx + 2, "Adjusted Quantity")

            df = df[cols]

        # ---------------- EXCEL ----------------
        wb = Workbook()

        # ORIGINAL
        ws0 = wb.active
        ws0.title = "Original Data"
        for r in dataframe_to_rows(original_df, index=False, header=True):
            ws0.append(r)

        # CLEANED
        ws1 = wb.create_sheet("Cleaned")
        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)

        # HEADER STYLING
        for i, col in enumerate(df.columns, 1):
            if col in ["Std Quantity", "Classification", "Item Price (USD)", 
                       "Standard Quantity", "Adjusted Quantity"]:
                ws1.cell(1, i).fill = light_orange
                ws1.cell(1, i).font = Font(bold=True)

        # VACCINE RED MARK
        if dataset_type == "vaccine":
            for row in range(2, ws1.max_row + 1):
                text = str(ws1.cell(row, df.columns.get_loc(desc)+1).value).lower()
                if "free sample" in text or "free quantity" in text:
                    ws1.cell(row, df.columns.get_loc(desc)+1).fill = dark_red

        # ---------------- ANALYSIS ----------------
        ws2 = wb.create_sheet("Analysis")

        def pivot(col, val):
            data = df.groupby(col)[val].sum().reset_index()
            return data.sort_values(by=val, ascending=False).head(10)

        tables = [
            ("Country vs Value (USD)", pivot(country, "Item Price (USD)")),
            ("Country vs Quantity", pivot(country, qty)),
            ("Consignee vs Value (USD)", pivot(consignee, "Item Price (USD)")),
            ("Exporter vs Value (USD)", pivot(exporter, "Item Price (USD)")),
            ("Exporter vs Quantity", pivot(exporter, qty))
        ]

        row = 1
        for title, data in tables:
            ws2.cell(row, 1, title)

            start = row + 1

            for i, rdata in enumerate(dataframe_to_rows(data, index=False, header=True)):
                for j, val in enumerate(rdata):
                    ws2.cell(start + i, j + 1, val)

            max_row = start + len(data)

            chart = BarChart()
            chart.title = title

            data_ref = Reference(ws2, min_col=2, min_row=start, max_row=max_row)
            cat_ref = Reference(ws2, min_col=1, min_row=start+1, max_row=max_row)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)

            ws2.add_chart(chart, f"E{row}")

            row += 15

        # ---------------- DASHBOARD ----------------
        ws3 = wb.create_sheet("Dashboard")

        ws3["A1"] = "DATASET DASHBOARD"
        ws3["A3"] = "Total Value (USD)"
        ws3["B3"] = df["Item Price (USD)"].sum()

        ws3["A4"] = "Total Quantity"
        ws3["B4"] = df[qty].sum()

        ws3["A5"] = "Top Country"
        ws3["B5"] = df.groupby(country)["Item Price (USD)"].sum().idxmax()

        wb.save("final.xlsx")

        st.success("✅ File Processed Successfully")

        with open("final.xlsx", "rb") as f:
            st.download_button("⬇ Download Excel", f, "final.xlsx")

    except Exception as e:
        st.error(f"❌ Error: {e}")