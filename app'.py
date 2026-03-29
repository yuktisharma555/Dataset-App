import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# ---------------- UI ----------------
st.set_page_config(layout="wide")

st.markdown("""
<style>
body {
    background: linear-gradient(135deg, #dbeafe, #ffffff, #ede9fe);
}
h1 {text-align:center; color:#6b21a8;}
</style>
""", unsafe_allow_html=True)

st.markdown("<h1>🦄 Smart Dataset Cleaner Pro</h1>", unsafe_allow_html=True)

file = st.file_uploader("Upload Excel", type=["xlsx"])
dataset_type = st.selectbox("Dataset Type", ["medicine","vaccine","testkit"])

# ---------------- COLORS ----------------
dark_red = PatternFill("solid", fgColor="8B0000")
light_red = PatternFill("solid", fgColor="FF9999")
light_orange = PatternFill("solid", fgColor="FFD580")

# ---------------- MAIN ----------------
if file:
    try:
        # ✅ FIXED HEADER ROW
        df = pd.read_excel(file, header=5)
        df.columns = df.columns.astype(str).str.upper().str.strip()

        st.success("✅ Header loaded from row 6 correctly")

        # -------- COLUMN MAPPING --------
        desc = "GOODS DESCRIPTION"
        qty = "QUANTITY"
        unit = "UNIT"
        price = "ITEM PRICE_INV"
        total_price = "TOTAL PRICE_INV_FC"
        country = "COUNTRY"
        exporter = "EXPORTER"
        consignee = "CONSIGNEE NAME"
        std_qty_col = "STD QUANTITY"

        # -------- SAFETY CHECK --------
        required = [desc, qty, price, country]
        missing = [c for c in required if c not in df.columns]

        if missing:
            st.error(f"❌ Missing columns: {missing}")
            st.stop()

        # -------- CLEAN TYPES --------
        df[qty] = pd.to_numeric(df[qty], errors="coerce")
        df[price] = pd.to_numeric(df[price], errors="coerce")
        df[total_price] = pd.to_numeric(df[total_price], errors="coerce")

        # -------- STANDARD QUANTITY --------
        if std_qty_col in df.columns:
            df["Std Quantity"] = pd.to_numeric(df[std_qty_col], errors="coerce")
        else:
            df["Std Quantity"] = df[qty]

        # -------- UNIT PRICE --------
        df["Unit Price"] = df[total_price] / df["Std Quantity"]

        # -------- CLASSIFICATION --------
        def classify(d):
            d = str(d).lower()
            if dataset_type == "vaccine":
                return "Pediatric" if "pediatric" in d else "Adult"
            return "Other"

        df["Classification"] = df[desc].apply(classify)

        # -------- CREATE EXCEL --------
        wb = Workbook()

        # ORIGINAL SHEET
        ws1 = wb.active
        ws1.title = "Original"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)

        # CLEANED SHEET
        ws2 = wb.create_sheet("Cleaned")

        for r in dataframe_to_rows(df, index=False, header=True):
            ws2.append(r)

        # -------- COLOR NEW HEADERS --------
        for i, col in enumerate(df.columns, 1):
            if col in ["Std Quantity", "Unit Price", "Classification"]:
                ws2.cell(1, i).fill = light_orange
                ws2.cell(1, i).font = Font(bold=True)

        desc_idx = list(df.columns).index(desc) + 1

        # -------- ROW COLOR RULES --------
        for i, row in df.iterrows():
            r = i + 2
            d = str(row[desc]).lower()

            if "free of cost" in d or "foc" in d:
                ws2.cell(r, desc_idx).fill = dark_red

            if pd.notna(row[qty]) and row[qty] < 1:
                ws2.cell(r, desc_idx).fill = dark_red

            if pd.notna(row[price]) and row[price] == 0:
                ws2.cell(r, desc_idx).fill = dark_red

            if any(x in d for x in ["sample","standard","r&d","impurity"]):
                ws2.cell(r, desc_idx).fill = dark_red

            if dataset_type == "medicine" and "api" in str(row["Classification"]).lower():
                ws2.cell(r, desc_idx).fill = light_red

        # -------- ANALYSIS --------
        ws3 = wb.create_sheet("Analysis")

        def pivot(col, val):
            return df.groupby(col)[val].sum().nlargest(10).reset_index()

        tables = [
            ("Country vs Unit Price", pivot(country, "Unit Price")),
            ("Country vs Quantity", pivot(country, "Std Quantity")),
            ("Consignee vs Unit Price", pivot(consignee, "Unit Price")),
            ("Exporter vs Unit Price", pivot(exporter, "Unit Price")),
            ("Exporter vs Quantity", pivot(exporter, "Std Quantity")),
            ("Classification vs Quantity", pivot("Classification", "Std Quantity"))
        ]

        row = 1
        for title, data in tables:
            ws3.cell(row, 1, title)

            for i, rdata in enumerate(dataframe_to_rows(data, index=False, header=True)):
                for j, val in enumerate(rdata):
                    ws3.cell(row + i + 1, j + 1, val)

            chart = BarChart()
            chart.title = title

            data_ref = Reference(ws3, min_col=2, min_row=row+2, max_row=row+11)
            cat_ref = Reference(ws3, min_col=1, min_row=row+3, max_row=row+11)

            chart.add_data(data_ref)
            chart.set_categories(cat_ref)

            ws3.add_chart(chart, f"E{row}")

            row += 15

        # -------- DASHBOARD --------
        ws4 = wb.create_sheet("Dashboard")

        ws4["A1"] = "DATASET DASHBOARD"
        ws4["A3"] = "Total Value"
        ws4["B3"] = df[total_price].sum()

        ws4["A4"] = "Total Quantity"
        ws4["B4"] = df["Std Quantity"].sum()

        ws4["A5"] = "Top Country"
        ws4["B5"] = tables[0][1].iloc[0,0] if not tables[0][1].empty else "N/A"

        # SAVE
        wb.save("final.xlsx")

        st.success("✅ Fully Processed Successfully")

        with open("final.xlsx", "rb") as f:
            st.download_button("⬇ Download Excel", f, "final.xlsx")

    except Exception as e:
        st.error(str(e))